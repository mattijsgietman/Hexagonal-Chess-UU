from openpyxl import load_workbook
import random

from Hex import Hex
from CONST import *
from Piece import *

class HexBoard():
    """
    Represents a hexagonal chess board.

    Attributes:
    - hexboard: A 2D list representing the hexagonal chess board.
    - random_puzzle: An integer representing the randomly generated puzzle number.

    Methods:
    - __init__(): Initializes the HexBoard object.
    - _create_board(): Creates the hexagonal chess board.
    - _setup_pieces(): Sets up the initial positions of the chess pieces.
    - get_piece(row, col): Returns the chess piece at the specified position.
    - set_piece(row, col, piece): Sets the chess piece at the specified position.
    - get_hexagon(row, col): Returns the hexagon object at the specified position.
    - get_pieces_locations(color): Returns a list of locations of the chess pieces of the specified color.
    - get_king_location(color): Returns the location of the king of the specified color.
    - get_pseudo_legal_moves(color): Returns a list of pseudo-legal moves for the chess pieces of the specified color.
    - get_legal_moves(color): Returns a list of legal moves for the chess pieces of the specified color.
    - move_piece(move, final=False): Moves a chess piece to the target position.
    - undo_move(move): Undoes a move by restoring the initial position of the chess piece.
    - in_check(color): Checks if the king of the specified color is in check.
    - is_game_over(color): Checks if the game is over for the specified color.
    - print_hexboard(): Prints the current state of the hexagonal chess board.
    - load_puzzle(puzzle): Loads a puzzle from an Excel file and sets up the chess pieces accordingly.
    """
    def __init__(self):
        """
        Initializes a new instance of the HexBoard class.
        """
        self.hexboard = [[None for _ in range(11)] for _ in range(21)]
        self._create_board()
        self._setup_pieces()
        self.random_puzzle = random.randint(1, 12)
        self.load_puzzle(str(self.random_puzzle))

    def _create_board(self):
        """
        Creates the initial hexagonal board by iterating over predefined positions.
        Each position is assigned a Hex object and added to the hexboard.
        """
        for position in POSITIONS:
            row, col = position
            hexagon = Hex(row, col, None)
            self.hexboard[row][col] = hexagon
    
    def _setup_pieces(self):
        """
        Set up the initial pieces on the hexagonal chessboard.

        This method assigns the appropriate pieces to their initial positions on the hexagonal chessboard.
        The black pieces are placed on the top half of the board, while the white pieces are placed on the bottom half.

        Note: This method assumes that the `hexboard` attribute has already been initialized.

        Returns:
            None
        """
        self.hexboard[0][5].piece = Bishop('black')
        self.hexboard[1][4].piece = Queen('black')
        self.hexboard[1][6].piece = King('black')
        self.hexboard[2][3].piece = Knight('black')
        self.hexboard[2][5].piece = Bishop('black')
        self.hexboard[2][7].piece = Knight('black')
        self.hexboard[3][2].piece = Rook('black')
        self.hexboard[3][8].piece = Rook('black')
        self.hexboard[4][1].piece = Pawn('black')
        self.hexboard[4][5].piece = Bishop('black')
        self.hexboard[4][9].piece = Pawn('black')
        self.hexboard[5][2].piece = Pawn('black')
        self.hexboard[5][8].piece = Pawn('black')
        self.hexboard[6][3].piece = Pawn('black')
        self.hexboard[6][7].piece = Pawn('black')
        self.hexboard[7][4].piece = Pawn('black')
        self.hexboard[7][6].piece = Pawn('black')
        self.hexboard[8][5].piece = Pawn('black')

        self.hexboard[20][5].piece = Bishop('white')
        self.hexboard[19][4].piece = Queen('white')
        self.hexboard[19][6].piece = King('white')
        self.hexboard[18][3].piece = Knight('white')
        self.hexboard[18][5].piece = Bishop('white')
        self.hexboard[18][7].piece = Knight('white')
        self.hexboard[17][2].piece = Rook('white')
        self.hexboard[17][8].piece = Rook('white')
        self.hexboard[16][1].piece = Pawn('white')
        self.hexboard[16][5].piece = Bishop('white')
        self.hexboard[16][9].piece = Pawn('white')
        self.hexboard[15][2].piece = Pawn('white')
        self.hexboard[15][8].piece = Pawn('white')
        self.hexboard[14][3].piece = Pawn('white')
        self.hexboard[14][7].piece = Pawn('white')
        self.hexboard[13][4].piece = Pawn('white')
        self.hexboard[13][6].piece = Pawn('white')
        self.hexboard[12][5].piece = Pawn('white')

    def get_piece(self, row, col):
        """
        Get the piece at the specified position on the hexagonal board.

        Args:
            row (int): The row index of the position.
            col (int): The column index of the position.

        Returns:
            object or None: The piece object at the specified position, or None if there is no piece.

        Raises:
            ValueError: If the position is invalid (outside the board boundaries).
        """
        if row < 0 or row > 20 or col < 0 or col > 10:
            raise ValueError("Invalid position")
        else:
            if self.hexboard[row][col] is not None:
                return self.hexboard[row][col].piece
            else:
                return None

    def set_piece(self, row, col, piece):
        """
        Sets a piece on the hexagonal board at the specified position.

        Args:
            row (int): The row index of the position.
            col (int): The column index of the position.
            piece (Piece): The piece to be set on the board.

        Raises:
            ValueError: If the position is invalid.

        """
        if row < 0 or row > 20 or col < 0 or col > 10:
            raise ValueError("Invalid position")
        else:
            self.hexboard[row][col].piece = piece
    
    def get_hexagon(self, row, col):
        """
        Retrieves the hexagon at the specified position.

        Args:
            row (int): The row index of the hexagon.
            col (int): The column index of the hexagon.

        Returns:
            object: The hexagon object at the specified position, or None if no hexagon exists.

        Raises:
            ValueError: If the position is invalid (row or col is out of bounds).
        """
        if row < 0 or row > 20 or col < 0 or col > 10:
            raise ValueError("Invalid position")
        elif self.hexboard[row][col] is not None:
            return self.hexboard[row][col]
        return None
    
    def get_pieces_locations(self, color):
        """
        Returns a list of locations (row, col) of all pieces of the specified color on the hexboard.

        Parameters:
        - color (str): The color of the pieces to search for.

        Returns:
        - list: A list of tuples representing the locations of the pieces.
        """
        locations = []
        for row in self.hexboard:
            for hexagon in row:
                if hexagon is not None:
                    if hexagon.piece is not None:
                        if hexagon.piece.color == color:
                            locations.append((hexagon.row, hexagon.col))
        return locations
    
    def get_king_location(self, color):
        """
        Returns the location of the king of the specified color on the hexagonal chessboard.

        Parameters:
        - color (str): The color of the king ('white' or 'black').

        Returns:
        - tuple: A tuple containing the row and column coordinates of the king's location.

        """
        for row in self.hexboard:
            for hexagon in row:
                if hexagon is not None:
                    if hexagon.piece is not None:
                        if hexagon.piece.color == color and isinstance(hexagon.piece, King):
                            return (hexagon.row, hexagon.col)

    def get_pseudo_legal_moves(self, color):
        """
        Returns a list of pseudo-legal moves for the specified color.

        Parameters:
        - color (str): The color of the pieces to consider.

        Returns:
        - list: A list of pseudo-legal moves for the specified color.

        """
        legal_moves = []
        locations = self.get_pieces_locations(color) 

        for location in locations:
            row, col = location
            piece = self.get_piece(row, col)
            legal_moves += piece._get_legal_moves(row, col, self)
        return legal_moves
    
    def get_legal_moves(self, color):
        """
        Returns a list of legal moves for the specified color.

        Parameters:
        - color (str): The color of the player whose legal moves are to be determined.

        Returns:
        - list: A list of legal moves for the specified color.

        """
        legal_moves = []
        moves = self.get_pseudo_legal_moves(color)
        for move in moves:
            self.move_piece(move)
            if not self.in_check(color):
                legal_moves.append(move)
            self.undo_move(move)
        return legal_moves
                
    def move_piece(self, move, final=False):
        """
        Moves a chess piece on the hexagonal chessboard.

        Args:
            move (Move): The move object containing the initial and target positions, as well as the piece and enemy_piece involved in the move.
            final (bool, optional): Indicates if this is the final move of the piece. Defaults to False.

        Returns:
            None

        Raises:
            None
        """
        initial = move.initial
        target = move.target
        piece = move.piece
        enemy_piece = move.enemy_piece
        initial_row, initial_col = initial
        target_row, target_col = target

        self.hexboard[initial_row][initial_col].piece = None
        self.hexboard[target_row][target_col].piece = piece

        if piece.name == 'p' or piece.name == 'P':
            if final:
                piece.has_moved = True
            if target in PAWN_PROMOTION_HEXAGONS:
                piece_index = self.hexboard[target_row][target_col].piece.index
                self.hexboard[target_row][target_col].piece = Queen(piece.color)
                self.hexboard[target_row][target_col].piece.index = piece_index
            piece.total_moves += 1

    def undo_move(self, move):
        """
        Undoes a move by restoring the initial state of the board.

        Args:
            move (Move): The move to be undone.

        Returns:
            None
        """
        initial = move.initial
        target = move.target
        piece = move.piece
        enemy_piece = move.enemy_piece
        initial_row, initial_col = initial
        target_row, target_col = target

        self.hexboard[initial_row][initial_col].piece = piece
        self.hexboard[target_row][target_col].piece = None

        if enemy_piece is not None:
            self.hexboard[target_row][target_col].piece = enemy_piece

        if piece.name == 'p' or piece.name == 'P':
            if target_row == 0 or target_row == 20:
                self.hexboard[initial_row][initial_col].piece = Pawn(piece.color)
            piece.total_moves -= 1

            if piece.total_moves == 0:
                piece.has_moved = False

    def in_check(self, color):
        """
        Checks if the specified color is in check.

        Args:
            color (str): The color of the player to check for check.

        Returns:
            bool: True if the specified color is in check, False otherwise.
        """
        king_location = self.get_king_location(color)

        opponent_color = 'black' if color == 'white' else 'white'
        opponent_locations = self.get_pieces_locations(opponent_color)

        for location in opponent_locations:
            piece = self.get_piece(*location)
            if piece is not None:
                legal_moves = piece._get_legal_moves(*location, self)
                for move in legal_moves:
                    if move.target == king_location:
                        return True

        return False
 
    def is_game_over(self, color):
        """
        Checks if the game is over for the specified color.

        Args:
            color (str): The color of the player to check for game over. Can be 'white' or 'black'.

        Returns:
            tuple: A tuple containing a boolean value indicating if the game is over, and a string indicating the result of the game.
                   If the game is over, the boolean value is True and the string can be 'black' (if white is in checkmate), 'white' (if black is in checkmate),
                   or 'remise' (if the game is a draw). If the game is not over, the boolean value is False and the string is an empty string.
        """
        if self.in_check('white'):
            legal_moves = self.get_legal_moves('white')
            if len(legal_moves) == 0:
                return (True, 'black')

        if self.in_check('black'):
            legal_moves = self.get_legal_moves('black')
            if len(legal_moves) == 0:
                return (True, 'white')
        
        if len(self.get_legal_moves(color)) == 0:
            return (True, 'remise')

        return (False, "")

    def print_hexboard(self):
        """
        Prints the current state of the hexboard.

        Each hexagon on the board is represented by a character.
        If a hexagon is empty, it is represented by a space character.
        If a hexagon has a piece on it, the piece's name is printed.
        If a hexagon has no piece, a dash character is printed.
        """
        for row in self.hexboard:
            for hexagon in row:
                if hexagon is None:
                    print(" ", end="")
                else:
                    if hexagon.piece is not None:
                        print(hexagon.piece.name, end="")
                    else:
                        print("-", end="")
            print()

    def load_puzzle(self, puzzle):
        """
        Loads a puzzle from an Excel file and populates the hexagonal chess board with the puzzle's pieces.

        Parameters:
        - puzzle (str): The name of the puzzle to load.

        Returns:
        - None
        """
        class_mapping = {
        'King': King,
        'Knight': Knight,
        'Bishop': Bishop,
        'Pawn': Pawn,
        'Queen': Queen,
        'Rook': Rook
        }
        self._create_board()
        filepath = f"puzzles/{puzzle}.xlsx"
        workbook = load_workbook(filepath, data_only=True)
        data = []
        king_counter = 0
        pawn_counter = 1
        knight_counter = 9
        bishop_counter = 11
        rook_counter = 14
        queen_counter = 16
        i=0
        for row in workbook.active.iter_rows(values_only=True):
            if i != 0:
                data.append(row)
            i+=1
        for row in data:
            piece, color, row, col, first_move = row
            self.hexboard[row][col].piece = class_mapping[piece](color)
            if piece == 'Pawn':
                if first_move == 'True':
                    self.hexboard[row][col].piece.has_moved = False
                    self.hexboard[row][col].piece.index = pawn_counter
                    pawn_counter += 1
                else:
                    self.hexboard[row][col].piece.has_moved = True
                    self.hexboard[row][col].piece.index = pawn_counter
                    pawn_counter += 1
            elif piece == 'Knight':
                self.hexboard[row][col].piece.index = knight_counter
                knight_counter += 1
            elif piece == 'Bishop':
                self.hexboard[row][col].piece.index = bishop_counter
                bishop_counter += 1
            elif piece == 'Rook':
                self.hexboard[row][col].piece.index = rook_counter
                rook_counter += 1
            elif piece == 'Queen':
                self.hexboard[row][col].piece.index = queen_counter
                queen_counter += 1
            elif piece == 'King':
                self.hexboard[row][col].piece.index = king_counter
                king_counter += 1

    def action_to_tuple(self, output):
        """
        Converts the output value to a tuple representing a piece and its position on the hexagonal board.

        Parameters:
        - output (int): The output value to be converted.

        Returns:
        - tuple: A tuple containing the piece and its position on the hexagonal board.
        """
        piece = output // 91
        hexagon = output % 91
        return (piece, POSITIONS[hexagon])
    
    def index_to_piece(self, index):
        """
        Returns the piece object with the given index.

        Parameters:
        index (int): The index of the piece to retrieve.

        Returns:
        Piece or None: The piece object with the given index, or None if no piece is found.

        """
        locations = self.get_pieces_locations('white') 
        for location in locations:
            row, col = location
            piece = self.get_piece(row, col)
            if piece != None:
                if piece.index == index:
                    return piece
    
    def legal_moves_to_actions(self, legal_moves):
        """
        Converts a list of legal moves to a list of actions.

        Args:
            legal_moves (list): A list of legal moves.

        Returns:
            list: A list of actions, where each action is a tuple containing the index of the piece and the target position.

        """
        actions = []
        for move in legal_moves:
            index = move.piece.index
            target = move.target
            actions.append((index, target))
        return actions
    
    def action_to_move(self, action):
        """
        Converts an action into a Move object.

        Parameters:
        - action (tuple): A tuple containing the index and target of the action.

        Returns:
        - Move: A Move object representing the converted action.

        """
        index, target = action
        locations = self.get_pieces_locations("white") 

        for location in locations:
            row, col = location
            piece = self.get_piece(row, col)
            if piece != None:
                if piece.index == index:
                    return Move(piece, location, target)
                
    def random_black_move(self):
        """
        Selects a random legal move for the black player.

        Returns:
            str or None: The selected move as a string, or None if no legal moves are available.
        """
        legal_moves = self.get_legal_moves("black")
        if len(legal_moves) == 0:
            return None
        move = random.choice(legal_moves)
        return move